#!/usr/bin/python3
#Lane Doyle
#12/12/19

'''create and populate an xlsx file called users.xlsx'''
import openpyxl
import datetime as dt
import time

try:
    print("Loading...")
    workbook = openpyxl.load_workbook("users.xlsx")
    sheet = workbook.get_sheet_by_name("Users")
    
    sheet['A1'].value = "Timestamp"
    sheet['B1'].value = "First Name"
    sheet['C1'].value = "Last Name"
    sheet['D1'].value = "Username"
    
    for i in range(2):
        new_row = sheet.max_row+1
        
        first = input("Enter first name of new user: ")
        last = input("Enter last name of new user: ")
        u_name = input("Enter username of new user: ")
        
        print("-----------------------")#Breaks up loop
        
        sheet.cell(row = new_row, column = 1).value = dt.datetime.now()
        sheet.cell(row = new_row, column = 2).value = first
        sheet.cell(row = new_row, column = 3).value = last
        sheet.cell(row = new_row, column = 4).value = u_name
        
    workbook.save("users.xlsx")
    print("Done!")    
    
except:
    print("Failed to load...creating new sheet!")
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Users")
    
    #Set the headings row
    sheet['A1'].value = "Timestamp"
    sheet['B1'].value = "First Name"
    sheet['C1'].value = "Last Name"
    sheet['D1'].value = "Username"
    
    for i in range(2):
        new_row = sheet.max_row+1
        
        first = input("Enter first name of new user: ")
        last = input("Enter last name of new user: ")
        u_name = input("Enter username of new user: ")
        
        print("-----------------------")#Breaks up loop
        
        sheet.cell(row = new_row, column = 1).value = dt.datetime.now()
        sheet.cell(row = new_row, column = 2).value = first
        sheet.cell(row = new_row, column = 3).value = last
        sheet.cell(row = new_row, column = 4).value = u_name
        
    workbook.save("users.xlsx")
    print("Done!")