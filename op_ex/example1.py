#!/usr/bin/python3
#Lane Doyle
#12/11/19

'''Program to demonstrate working with spreedsheets'''

import openpyxl

workbook = openpyxl.load_workbook('example1.xlsx')
sheet = workbook.get_sheet_by_name("Sheet1")
print(sheet.max_row)

fruits = []

for i in range(1,sheet.max_row + 1):
    fruits.append(sheet.cell(row = i, column = 2).value)
print(fruits)
fruits[0] = "Eepples"
print(fruits)
print(sheet['B1'].value)

for i in range(1,sheet.max_row + 1):
    sheet.cell(row = i, column = 2).value = fruits[i-1]
print(sheet['B1'].value)

print(sheet.max_row)
print(sheet.max_column)
new_col = sheet.max_column+1

for i in range (1, sheet.max_row + 1):
    sheet.cell(row = i, column = new_col).value = "eaten"

workbook.save('example1.xlsx')


    
    